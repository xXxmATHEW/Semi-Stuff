import openpyxl
a=0
b=2
while a==0:
    temp=int(input("Temperatur:"))
    konz=int(input("Konzentration:"))
    if konz==5:
             a=1
             print("Messvorgang abgeschlossen")
    else:
        fileXLSX = openpyxl.load_workbook("Auswertung.xlsx")
        sheet = fileXLSX["Tabelle1"]
        sheet.cell(row=b, column=1).value = temp/1000
        sheet.cell(row=b, column=2).value = konz/1000
        fileXLSX.save('Auswertung.xlsx')
        b=b+1
    



